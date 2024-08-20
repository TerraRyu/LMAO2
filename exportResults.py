import openpyxl
from openpyxl.styles import Font, Alignment
import re
from datetime import datetime
import io

def sanitize_filename(filename):
    return re.sub(r'[\\/*?:"<>|]', "", filename)

def export_to_excel(enumeration_results, domain, output):
    if not enumeration_results:
        return "No results to export."

    wb = openpyxl.Workbook()
    
    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)

    # Create sheets for each tab
    create_all_subdomains_sheet(wb, enumeration_results)
    create_virustotal_sheet(wb, enumeration_results)
    create_dnsdumpster_sheet(wb, enumeration_results)

    # Save to the provided output (BytesIO object)
    wb.save(output)
    return "Results exported successfully"

def create_all_subdomains_sheet(wb, results):
    ws = wb.create_sheet("All Subdomains")
    headers = ['Subdomain', 'IP Address', 'HTTP Status', 'Source']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row = 2
    for subdomain, info in results['subdomains'].items():
        ws.cell(row=row, column=1, value=subdomain)
        ws.cell(row=row, column=2, value=info.get('ip', 'N/A'))
        ws.cell(row=row, column=3, value=info.get('status', 'N/A'))
        ws.cell(row=row, column=4, value=', '.join(info.get('sources', [])))
        row += 1

    adjust_column_width(ws)

def create_virustotal_sheet(wb, results):
    ws = wb.create_sheet("VirusTotal")
    headers = ['Subdomain', 'IP Address']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row = 2
    for subdomain, ips in results['virustotal'].items():
        ws.cell(row=row, column=1, value=subdomain)
        ws.cell(row=row, column=2, value=', '.join(ips))
        row += 1

    adjust_column_width(ws)

def create_dnsdumpster_sheet(wb, results):
    ws = wb.create_sheet("DNSDumpster")
    
    # Subdomains
    ws.append(["Subdomains"])
    headers = ['Subdomain', 'IP Address', 'ASN', 'Server']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row = 3
    for subdomain in results['dnsdumpster'].get('subdomains', []):
        ws.cell(row=row, column=1, value=subdomain.get('domain', 'N/A'))
        ws.cell(row=row, column=2, value=subdomain.get('ip', 'N/A'))
        ws.cell(row=row, column=3, value=str(subdomain.get('asn', 'N/A')))
        ws.cell(row=row, column=4, value=subdomain.get('server', 'N/A'))
        row += 1

    row += 2  # Add some space

    # MX Records
    ws.cell(row=row, column=1, value="MX Records")
    row += 1
    headers = ['Exchange', 'IP Address', 'Preference', 'Source']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row += 1
    for mx_record in results['dnsdumpster'].get('mx_records', []):
        ws.cell(row=row, column=1, value=mx_record.get('exchange', 'N/A'))
        ws.cell(row=row, column=2, value=mx_record.get('ip', 'N/A'))
        ws.cell(row=row, column=3, value=mx_record.get('preference', 'N/A'))
        ws.cell(row=row, column=4, value=mx_record.get('source', 'N/A'))
        row += 1

    row += 2  # Add some space

    # TXT Records
    ws.cell(row=row, column=1, value="TXT Records")
    row += 1
    for txt_record in results['dnsdumpster'].get('txt_records', []):
        ws.cell(row=row, column=1, value=str(txt_record))
        row += 1

    row += 2  # Add some space

    # DNS Servers
    ws.cell(row=row, column=1, value="DNS Servers")
    row += 1
    headers = ['DNS Server', 'IP Address']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row += 1
    for dns_record in results['dnsdumpster'].get('dns_records', []):
        ws.cell(row=row, column=1, value=dns_record.get('ns', 'N/A'))
        ws.cell(row=row, column=2, value=dns_record.get('ip', 'N/A'))
        row += 1

    adjust_column_width(ws)

def adjust_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width